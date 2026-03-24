import re
from pathlib import Path
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ================= CONFIG =================
HORIZONTAL_BASE_DIR = Path("results/horizontal-scaling")
VERTICAL_BASE_DIR = Path("results/vertical-scaling")
OUT_DIR = Path("generated-results/output")
OUT_DIR.mkdir(parents=True, exist_ok=True)

OUT_FILE = OUT_DIR / "scaling_summary_full.xlsx"

DELTA = 4
TOTAL_TIME = 5

IMPLEMENTATIONS = [
    "python-monolith-images",
    "javascript-monolith-images",
    "java-monolith-images",
    "python-microservices-grpc-images",
    "javascript-microservices-grpc-images",
    "java-microservices-grpc-images",
]

HORIZONTAL_LEVELS = ["1-instance", "2-instance", "4-instance"]
VERTICAL_LEVELS = ["c2d-standard-2", "c2d-standard-4", "c2d-standard-8"]

LANGUAGE_MAP = {
    "python": "Python",
    "javascript": "JavaScript",
    "java": "Java",
}

ARCHITECTURE_MAP = {
    "monolith": "monolityczna",
    "microservices": "mikroserwisowa",
}

COMMUNICATION_MAP = {
    "grpc": "gRPC",
    "rest": "REST API",
    "none": "-",
}

# ================= HELPERS =================
def extract_users(fname: str) -> int:
    match = re.search(r"users_(\d+)", fname)
    if not match:
        raise ValueError(f"Nie można odczytać liczby użytkowników z nazwy pliku: {fname}")
    return int(match.group(1))


def get_benchmark_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["response_end"] = df["timeStamp"] + df["elapsed"]

    start_est = df["timeStamp"].min() + DELTA * 1000
    valid_after_ramp = df[df["response_end"] > start_est]

    if valid_after_ramp.empty:
        return df.iloc[0:0]

    start = valid_after_ramp["response_end"].min()
    end = start + TOTAL_TIME * 1000

    return df[(df["response_end"] >= start) & (df["response_end"] <= end)]


def compute_throughput(df: pd.DataFrame) -> float:
    if df.empty:
        return 0.0
    df = df[df["responseCode"] == 200]
    return len(df) / TOTAL_TIME


def parse_implementation(implementation: str) -> tuple[str, str, str]:
    if implementation.endswith("-images"):
        implementation = implementation[:-7]

    parts = implementation.split("-")
    language = parts[0]

    if "monolith" in parts:
        architecture = "monolith"
        communication = "none"
    elif "microservices" in parts and "grpc" in parts:
        architecture = "microservices"
        communication = "grpc"
    elif "microservices" in parts and "rest" in parts:
        architecture = "microservices"
        communication = "rest"
    else:
        raise ValueError(f"Nieobsługiwana implementacja: {implementation}")

    return language, architecture, communication


def percent_change(new_value, old_value):
    if old_value is None or pd.isna(old_value) or old_value == 0:
        return None
    if new_value is None or pd.isna(new_value):
        return None
    return ((new_value - old_value) / old_value) * 100.0


def percent_diff(a, b):
    if b is None or pd.isna(b) or b == 0:
        return None
    if a is None or pd.isna(a):
        return None
    return ((a - b) / b) * 100.0


def to_excel_percent(value):
    if value is None or pd.isna(value):
        return None
    return value / 100.0


def load_scaling_data(base_dir: Path, scaling_kind: str, levels: list[str], level_column_name: str) -> pd.DataFrame:
    rows = []

    if not base_dir.exists():
        return pd.DataFrame(columns=["implementation", level_column_name, "users", "throughput"])

    for level_dir in base_dir.iterdir():
        if not level_dir.is_dir():
            continue

        level_name = level_dir.name
        if level_name not in levels:
            continue

        for impl_dir in level_dir.iterdir():
            if not impl_dir.is_dir():
                continue

            implementation = impl_dir.name
            if implementation not in IMPLEMENTATIONS:
                continue

            for iteration_dir in impl_dir.iterdir():
                if not iteration_dir.is_dir():
                    continue

                for csv_file in iteration_dir.glob("users_*.csv"):
                    df = pd.read_csv(csv_file)
                    df["timeStamp"] -= df["timeStamp"].min()

                    bench = get_benchmark_df(df)
                    thr = compute_throughput(bench)

                    rows.append({
                        "implementation": implementation,
                        level_column_name: level_name,
                        "users": extract_users(csv_file.name),
                        "throughput": thr,
                        "scaling_kind": scaling_kind
                    })

    if not rows:
        return pd.DataFrame(columns=["implementation", level_column_name, "users", "throughput", "scaling_kind"])

    df = pd.DataFrame(rows)

    agg = (
        df.groupby(["implementation", level_column_name, "users", "scaling_kind"], as_index=False)
          .median(numeric_only=True)
    )

    return agg


def get_max_for_level(df: pd.DataFrame, implementation: str, level_column: str, level_value: str):
    subset = df[
        (df["implementation"] == implementation) &
        (df[level_column] == level_value)
    ]
    if subset.empty:
        return None
    return float(subset["throughput"].max())


# ================= BUILD TABLES =================
def build_horizontal_summary(horizontal_df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    lp = 1

    for implementation in IMPLEMENTATIONS:
        language_code, architecture_code, communication_code = parse_implementation(implementation)

        max_1 = get_max_for_level(horizontal_df, implementation, "instance", "1-instance")
        max_2 = get_max_for_level(horizontal_df, implementation, "instance", "2-instance")
        max_4 = get_max_for_level(horizontal_df, implementation, "instance", "4-instance")

        rows.append({
            "Lp.": lp,
            "Język programowania": LANGUAGE_MAP.get(language_code, language_code),
            "Architektura": ARCHITECTURE_MAP.get(architecture_code, architecture_code),
            "Technologia komunikacji": COMMUNICATION_MAP.get(communication_code, communication_code),
            "Maks. przepustowość 1 inst. [req/s]": max_1,
            "Maks. przepustowość 2 inst. [req/s]": max_2,
            "Maks. przepustowość 4 inst. [req/s]": max_4,
            "Wzrost 2 inst. vs 1 inst.": to_excel_percent(percent_change(max_2, max_1)),
            "Wzrost 4 inst. vs 1 inst.": to_excel_percent(percent_change(max_4, max_1)),
            "Wzrost 4 inst. vs 2 inst.": to_excel_percent(percent_change(max_4, max_2)),
            "implementation_key": implementation,
        })
        lp += 1

    return pd.DataFrame(rows)


def build_vertical_summary(vertical_df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    lp = 1

    for implementation in IMPLEMENTATIONS:
        language_code, architecture_code, communication_code = parse_implementation(implementation)

        max_2 = get_max_for_level(vertical_df, implementation, "machine", "c2d-standard-2")
        max_4 = get_max_for_level(vertical_df, implementation, "machine", "c2d-standard-4")
        max_8 = get_max_for_level(vertical_df, implementation, "machine", "c2d-standard-8")

        rows.append({
            "Lp.": lp,
            "Język programowania": LANGUAGE_MAP.get(language_code, language_code),
            "Architektura": ARCHITECTURE_MAP.get(architecture_code, architecture_code),
            "Technologia komunikacji": COMMUNICATION_MAP.get(communication_code, communication_code),
            "Maks. przepustowość std-2 [req/s]": max_2,
            "Maks. przepustowość std-4 [req/s]": max_4,
            "Maks. przepustowość std-8 [req/s]": max_8,
            "Wzrost std-4 vs std-2": to_excel_percent(percent_change(max_4, max_2)),
            "Wzrost std-8 vs std-2": to_excel_percent(percent_change(max_8, max_2)),
            "Wzrost std-8 vs std-4": to_excel_percent(percent_change(max_8, max_4)),
            "implementation_key": implementation,
        })
        lp += 1

    return pd.DataFrame(rows)


def build_comparison_summary(horizontal_summary: pd.DataFrame, vertical_summary: pd.DataFrame) -> pd.DataFrame:
    merged = horizontal_summary.merge(
        vertical_summary,
        on="implementation_key",
        suffixes=("_h", "_v")
    )

    rows = []
    lp = 1

    for _, row in merged.iterrows():
        h2 = row["Maks. przepustowość 2 inst. [req/s]"]
        h4 = row["Maks. przepustowość 4 inst. [req/s]"]
        v2 = row["Maks. przepustowość std-4 [req/s]"]
        v4 = row["Maks. przepustowość std-8 [req/s]"]

        rows.append({
            "Lp.": lp,
            "Język programowania": row["Język programowania_h"],
            "Architektura": row["Architektura_h"],
            "Technologia komunikacji": row["Technologia komunikacji_h"],
            "Horyzontalne 2x [req/s]": h2,
            "Wertykalne 2x [req/s]": v2,
            "Przewaga vertical 2x nad horizontal 2x": to_excel_percent(percent_diff(v2, h2)),
            "Horyzontalne 4x [req/s]": h4,
            "Wertykalne 4x [req/s]": v4,
            "Przewaga vertical 4x nad horizontal 4x": to_excel_percent(percent_diff(v4, h4)),
        })
        lp += 1

    return pd.DataFrame(rows)


# ================= EXCEL FORMATTING =================
def format_sheet(ws, percent_columns=None, throughput_columns=None):
    percent_columns = percent_columns or []
    throughput_columns = throughput_columns or []

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            if cell.column == 1:
                cell.alignment = center
            else:
                cell.alignment = wrap

    for col_letter in throughput_columns:
        for cell in ws[col_letter][1:]:
            cell.number_format = "0.0"

    for col_letter in percent_columns:
        for cell in ws[col_letter][1:]:
            cell.number_format = "0.0%"

    widths = {
        "A": 6,
        "B": 22,
        "C": 18,
        "D": 24,
        "E": 22,
        "F": 22,
        "G": 22,
        "H": 22,
        "I": 22,
        "J": 24,
        "K": 18,
    }
    for col, width in widths.items():
        if col in ws.column_dimensions:
            ws.column_dimensions[col].width = width
        else:
            ws.column_dimensions[col].width = width


# ================= MAIN =================
def main():
    horizontal_df = load_scaling_data(
        base_dir=HORIZONTAL_BASE_DIR,
        scaling_kind="horizontal",
        levels=HORIZONTAL_LEVELS,
        level_column_name="instance"
    )

    vertical_df = load_scaling_data(
        base_dir=VERTICAL_BASE_DIR,
        scaling_kind="vertical",
        levels=VERTICAL_LEVELS,
        level_column_name="machine"
    )

    if horizontal_df.empty and vertical_df.empty:
        raise RuntimeError("Brak danych w results/horizontal-scaling i results/vertical-scaling.")

    horizontal_summary = build_horizontal_summary(horizontal_df)
    vertical_summary = build_vertical_summary(vertical_df)
    comparison_summary = build_comparison_summary(horizontal_summary, vertical_summary)

    horizontal_to_save = horizontal_summary.drop(columns=["implementation_key"])
    vertical_to_save = vertical_summary.drop(columns=["implementation_key"])

    with pd.ExcelWriter(OUT_FILE, engine="openpyxl") as writer:
        horizontal_to_save.to_excel(writer, sheet_name="Horizontal", index=False)
        vertical_to_save.to_excel(writer, sheet_name="Vertical", index=False)
        comparison_summary.to_excel(writer, sheet_name="Horizontal_vs_Vertical", index=False)

        wb = writer.book

        ws_h = writer.sheets["Horizontal"]
        format_sheet(
            ws_h,
            percent_columns=["H", "I", "J"],
            throughput_columns=["E", "F", "G"]
        )

        ws_v = writer.sheets["Vertical"]
        format_sheet(
            ws_v,
            percent_columns=["H", "I", "J"],
            throughput_columns=["E", "F", "G"]
        )

        ws_cmp = writer.sheets["Horizontal_vs_Vertical"]
        format_sheet(
            ws_cmp,
            percent_columns=["G", "J"],
            throughput_columns=["E", "F", "H", "I"]
        )

    print(f"Gotowe. Zapisano plik: {OUT_FILE}")


if __name__ == "__main__":
    main()